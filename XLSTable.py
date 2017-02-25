"""
Access worksheets of excel files as tables with query-oriented named columns.
At the same time we preserve the usability when interactively working
in a spreadsheet application.

WARNING: Don't have excel open, while you write to excel files from python

A table may be placed anywhere on the work sheet, as long as the zero-th row
provides a 'table descriptor' based on tags:

  A B  C          D              E              F              G              H ...
1 .
2   .... these lines are ignored and can be used for comments and other header stuff.
3   .................
4      __TAB__    ColumnName1    ColumnName1    ColumnName1    ColumnName1    __RIGHT__
5                 X <---first table row starts here and is described as 'D5:G5' in spreadsheet terms

The cell containing the tag __TAB__ is the upper left corner of the table, followed by the titles of the columns. Make sure they are unique!
__RIGHT__ and __BOTTOM__ may be ommitted but mark the right/bottom edge of the table.
"""

import openpyxl
from collections import OrderedDict
from openpyxl.styles.colors import Color


### Post-Modify the Cell class so that it has an additional 'evaluate' method plugged into the existing API / library
def evaluateCell(cell,default=None):
	""" Returns the value of a cell.
	`load_datamirror_workbook` will bind `evaluateCell` as a method `evaluate` to the :class:`.Cell`.

	.. todo:: yes, I know I could/should? use a decorator for this

	`somecellinstance.evaluate()` will give transparent access to the value of the cell (irrespective of whether it is defined as a formula or as a value in the workbook).
	If default != None, then any empty cells will be replaced by that default.


	"""
	if cell.data_type == 'f':
		sheettitle = cell.parent.title
		result = cell.parent.parent.data_only_mirror[sheettitle][cell.coordinate].value
	else:
		result = cell.value
	if result == None: result=default
	return result

#
#TODO: .
#one after the other
openpyxl.cell.cell.Cell.evaluate = evaluateCell

def setValueCell(cell,value):
	""" Sets the value of a cell. Style will be copied, the font color will be green to indicate the machine generation of the value."""
	fnt = cell.font.copy(color=Color(rgb="00604361"))
	fll = cell.fill.copy()
	cell.value = value
	cell.font = fnt
	cell.fill = fll

openpyxl.cell.cell.Cell.setValue = setValueCell 	#bind evaluateCell as a method to the class Cell

def load_datamirror_workbook(filename):
	""" Load a workbook in twin mode to have access on formulae and values.

	openpyxl requires to use a workbook either based on the fomulae OR based on the resulting data/values.

	This method loads two twins of the workbook:
	1. for reading and writing, access to formulae.
	2. a value mirror (read-only), for accessing cell values instead of formulae. (read-only, writing to it would make openpyxl turn all the formulae in the workbook into values, even in cells which we did not touch at all.)
	"""
	wb 			= openpyxl.load_workbook(filename,data_only=False)
	wb_mirror 	= openpyxl.load_workbook(filename,data_only=True,read_only=True)
	wb.data_only_mirror = wb_mirror
	return wb

class XLSTable(object):
	TAB_DESCRIPTOR = u"__TAB__"
	TAB_RIGHTBOUND = u"__RIGHT__"
	TAB_BOTTOMBOUND = u"__BOTTOM__"
	def __init__(self,worksheet):
		self.ws = worksheet
		self.initialMaxCol = self.ws.max_column
		self.initialMaxRow = self.ws.max_row
		self.columns = OrderedDict()
		self.firstrow = 0
		self.descr = self.getTableDescriptor()
		self.detectColumns()
		self.rowsN = self.getRowsN()
		#self.rows = []

	def getTableDescriptor(self):
		""" Returns the cell containing '__TAB__'.
			The reminder of the row contains the column names.
			'__RICHT__' should mark the right edge of the table descriptor row."""
		for row in self.ws:
			for cell in row:
				if cell.value == self.TAB_DESCRIPTOR:
					return cell

		raise ValueError("Could not detect Table Descriptor Line. There should be a cell containg " + self.TAB_DESCRIPTOR)

	def detectColumns(self):
		c = self.descr.offset(column=1)
		#print ("column max:", self.ws.max_column)
		while c.col_idx <= self.initialMaxCol and c.value != self.TAB_RIGHTBOUND:
			#print (c.col_idx,c)
			self.columns[c.value] = c
			c = c.offset(column=1)

	def getRowsN(self):
		""" How many Rows .... does this table have....before you can see the end? """
		c = self.descr.offset(row=1)
		rowsN = 0
		while c.row <= self.initialMaxRow and c.value != self.TAB_BOTTOMBOUND:
			c = c.offset(row=1)
			rowsN +=1
		return rowsN

	def getColumn ( self, key, filterFn=(lambda x:x) ):
		""" Returns an iterator over all cells in the column identified by key."""
		column,r1 = self.columns[key].column,self.columns[key].row + 1
		rn = r1 + self.rowsN - 1
		return filter(filterFn,map(lambda el:el[0], self.ws[column+str(r1)+':'+column+str(rn)]) )

	def iterRows(self, pattern=None):
		""" Iterate over all rows. `pattern` is used to filter the rows. The keys define the columns to use for the filtering, the corresponding
			values define the exact values the cells should evaluate to.
			:param pattern: A dictionary where each `key`/`value` pair specifies a requirement so that the respective cell in the `key`-column evalueates to the corresponding `value`.
			:return:
			"""
		if pattern != None:
			return FilteredRowsIterator(self,pattern)
		else:
			return RowsIterator(self)

	def getRowAsDict(self,rowOfTable):
		""" Returns an :Class:`OrderedDict` containing the cells in the row.
			:param rowOfTable: offset in the table from the descriptor line. (i.e. 1 means first data row)
		"""
		therow = OrderedDict()
		for key,cell in self.columns.items():
			c = cell.offset(row=rowOfTable)
			therow[key] = c
		return therow

	def getRowOfCell(self,cell):
		therow = OrderedDict()
		row = str(cell.row)
		for key,headercell in self.columns.items():
			c = self.ws[headercell.column+row]
			therow[key] = c
		return therow

class RowsIterator(object):
	def __init__(self,table):
		self.countdown = table.rowsN
		self.cell = table.descr
		self.table = table

	def __iter__(self):
		return self

	def __next__(self):
		if self.countdown == 0: raise StopIteration()
		else:
			self.countdown-=1
			self.cell = self.cell.offset(row=1)
			return self.table.getRowOfCell(self.cell)

class FilteredRowsIterator(object):
	""" Creates an iterator over all rows, where the rows are filtered according to a pattern.

		Technically, the pattern is a dictionary defining the columns to use for the filtering based on its keys and the exact values each row must evaluate to."""

	cells2tuple = lambda line:tuple(map ( lambda c:c.value, line ))	# go through a series of cells and put the values in a tuble

	def __init__(self,table,pattern):

		### retrieve
		self.thetable = table

		values = tuple ( pattern.values() )   													### put the values of the pattern into a tuple
		extractValue = lambda cellstuple: tuple(map ( lambda cell:cell.value, cellstuple ))		### unpack, evaluate each Row-tuple
		filterFn = lambda cellstuple:extractValue(cellstuple) == values							### check each row tuple to be equal to the values tuple

		### create a filter that iterates over all rows in the table
		### and zips into tuples all columns that are relevant for the pattern
		### (Cell<Row1/Patternkey1>,Cell<Row1/Patternkey2>,...) , (Cell<Row2/Patternkey1>,Cell<Row2/Patternkey2>,...)
		self.thefilter = filter ( filterFn , zip(*[self.thetable.getColumn(key) for key in pattern.keys()]) )

	def __iter__(self):
		return self

	def __next__(self):
		nextitem = self.thefilter.__next__()
		return self.thetable.getRowOfCell(nextitem[0])

def main():
	"""Quick demo, how to use the module."""
	wb = load_datamirror_workbook('example.xlsx')
	tab = XLSTable(wb['Sheet1'])
	print ("Iterate over all rows.")
	for itemdict in tab.iterRows():
		print (itemdict['Name'].value,itemdict['Age'].value,itemdict['Age'].evaluate())

	print ("")
	print ("Iterate over all rows with parameter pattern={'ZIP Code':4469,'Age':40}")
	for itemdict in tab.iterRows(pattern={'ZIP Code':4469,'Age':40}):
		print (itemdict['Name'].value,itemdict['Age'].value,itemdict['Age'].evaluate())


if __name__ == "__main__":
    main()
